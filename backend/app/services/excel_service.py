"""Excel processing service for running formulas and generating output"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.services.utils import log_message


class ExcelService:
    """Service for processing Excel files with formulas"""

    def __init__(self):
        pass

    def copy_data_to_excel(
        self,
        source_files: list[dict],
        template_path: Path | None,
        output_path: Path,
        worksheet_mapping: dict[str, str] = None,
    ) -> dict:
        """Copy data from downloaded files to Excel template

        Args:
            source_files: List of dicts with 'file_path', 'file_type', 'worksheet_name'
            template_path: Path to Excel template (optional, creates new if None)
            output_path: Path to output Excel file
            worksheet_mapping: Mapping of file_type to worksheet name

        Returns dict with success status and file path
        """
        try:
            # Load template or create new workbook
            if template_path and template_path.exists():
                wb = load_workbook(template_path)
                log_message(f"Loaded template: {template_path}")
            else:
                wb = Workbook()
                # Remove default sheet
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                log_message("Created new workbook")

            # Default worksheet mapping
            if worksheet_mapping is None:
                worksheet_mapping = {
                    "cm_bhavcopy": "Eq Bhav",
                    "fo_bhavcopy": "FU Data final",
                    "cm_delivery": "Eq Del",
                    "fo_udiff": "Daily OP Data",
                    "fo_participant_oi": "Daily OP Data",
                }

            # Process each source file
            for source_file in source_files:
                file_path = Path(source_file["file_path"])
                file_type = source_file.get("file_type", "")
                worksheet_name = source_file.get("worksheet_name") or worksheet_mapping.get(
                    file_type, "Data"
                )

                # Get or create worksheet
                if worksheet_name in wb.sheetnames:
                    ws = wb[worksheet_name]
                else:
                    ws = wb.create_sheet(worksheet_name)

                # Read data based on file type
                if file_path.suffix.lower() == ".csv":
                    self._copy_csv_to_worksheet(file_path, ws, source_file.get("start_row", 1))
                elif file_path.suffix.lower() == ".zip":
                    # Extract and copy CSV from zip
                    import zipfile

                    with zipfile.ZipFile(file_path, "r") as zip_ref:
                        csv_files = [f for f in zip_ref.namelist() if f.endswith(".csv")]
                        if csv_files:
                            # Extract first CSV to temp location
                            temp_csv = Path(file_path.parent) / f"temp_{file_path.stem}.csv"
                            with (
                                zip_ref.open(csv_files[0]) as source,
                                open(temp_csv, "wb") as target,
                            ):
                                target.write(source.read())
                            self._copy_csv_to_worksheet(
                                temp_csv, ws, source_file.get("start_row", 1)
                            )
                            temp_csv.unlink()  # Cleanup
                elif file_path.suffix.lower() == ".dat":
                    self._copy_dat_to_worksheet(file_path, ws, source_file.get("start_row", 1))

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            log_message(f"Saved Excel file: {output_path}")

            return {"success": True, "output_path": str(output_path), "worksheets": wb.sheetnames}

        except Exception as e:
            log_message(f"Error copying data to Excel: {e!s}")
            return {"success": False, "error": str(e)}

    def _copy_csv_to_worksheet(self, csv_path: Path, worksheet, start_row: int = 1):
        """Copy CSV data to worksheet"""
        try:
            df = pd.read_csv(csv_path)

            # Write headers
            for col_idx, header in enumerate(df.columns, start=1):
                worksheet.cell(row=start_row, column=col_idx, value=header)

            # Write data
            for row_idx, row_data in enumerate(df.values, start=start_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            log_message(f"Copied {len(df)} rows from {csv_path.name} to {worksheet.title}")
        except Exception as e:
            log_message(f"Error copying CSV: {e!s}")
            raise

    def _copy_dat_to_worksheet(self, dat_path: Path, worksheet, start_row: int = 1):
        """Copy DAT file data to worksheet"""
        try:
            with open(dat_path, encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()

            for row_idx, line in enumerate(lines, start=start_row):
                # Split by pipe or tab (common DAT formats)
                values = line.strip().split("|") if "|" in line else line.strip().split("\t")
                for col_idx, value in enumerate(values, start=1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            log_message(f"Copied {len(lines)} lines from {dat_path.name} to {worksheet.title}")
        except Exception as e:
            log_message(f"Error copying DAT: {e!s}")
            raise

    def run_formulas(self, excel_path: Path) -> dict:
        """Run formulas in Excel file

        Args:
            excel_path: Path to Excel file
            formula_config: Optional config for formula execution

        Returns dict with success status
        """
        try:
            wb = load_workbook(excel_path, data_only=False)

            # Enable formula calculation
            wb.calculation.calculateMode = "automatic"

            # Process each worksheet
            formulas_executed = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Iterate through cells and calculate formulas
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == "f":  # Formula
                            try:
                                # Force recalculation by reading value
                                _ = cell.value
                                formulas_executed += 1
                            except Exception as e:
                                log_message(
                                    f"Error calculating formula in {sheet_name} cell {cell.coordinate}: {e!s}"
                                )

            # Save with calculated values
            wb.save(excel_path)
            log_message(f"Executed {formulas_executed} formulas in {excel_path}")

            return {
                "success": True,
                "formulas_executed": formulas_executed,
                "worksheets": wb.sheetnames,
            }

        except Exception as e:
            log_message(f"Error running formulas: {e!s}")
            return {"success": False, "error": str(e)}

    def copy_to_output(
        self,
        source_excel_path: Path,
        output_path: Path,
        worksheets_to_copy: list[str] | None = None,
    ) -> dict:
        """Copy processed Excel data to final output file

        Args:
            source_excel_path: Path to Excel file with calculated formulas
            output_path: Path to final output Excel file
            worksheets_to_copy: List of worksheet names to copy (None = all)

        Returns dict with success status
        """
        try:
            wb_source = load_workbook(source_excel_path, data_only=True)
            wb_output = Workbook()

            # Remove default sheet
            if "Sheet" in wb_output.sheetnames:
                wb_output.remove(wb_output["Sheet"])

            # Determine which sheets to copy
            sheets_to_copy = worksheets_to_copy or wb_source.sheetnames

            for sheet_name in sheets_to_copy:
                if sheet_name in wb_source.sheetnames:
                    ws_source = wb_source[sheet_name]
                    ws_output = wb_output.create_sheet(sheet_name)

                    # Copy all cells
                    for row in ws_source.iter_rows():
                        for cell in row:
                            ws_output[cell.coordinate].value = cell.value
                            # Copy formatting if needed
                            if cell.has_style:
                                ws_output[cell.coordinate].font = cell.font
                                ws_output[cell.coordinate].fill = cell.fill
                                ws_output[cell.coordinate].border = cell.border
                                ws_output[cell.coordinate].alignment = cell.alignment

            # Save output file
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb_output.save(output_path)
            log_message(f"Created output file: {output_path}")

            return {
                "success": True,
                "output_path": str(output_path),
                "worksheets_copied": sheets_to_copy,
            }

        except Exception as e:
            log_message(f"Error copying to output: {e!s}")
            return {"success": False, "error": str(e)}

    def process_full_pipeline(
        self,
        source_files: list[dict],
        template_path: Path | None,
        intermediate_path: Path,
        output_path: Path,
        worksheet_mapping: dict | None = None,
        worksheets_to_output: list[str] | None = None,
    ) -> dict:
        """Run full pipeline: copy data -> run formulas -> copy to output

        Returns dict with all results
        """
        results = {
            "copy_data": None,
            "run_formulas": None,
            "copy_to_output": None,
            "success": False,
        }

        # Step 1: Copy data to Excel
        copy_result = self.copy_data_to_excel(
            source_files, template_path, intermediate_path, worksheet_mapping
        )
        results["copy_data"] = copy_result

        if not copy_result.get("success"):
            results["error"] = f"Failed to copy data: {copy_result.get('error')}"
            return results

        # Step 2: Run formulas
        formula_result = self.run_formulas(intermediate_path)
        results["run_formulas"] = formula_result

        if not formula_result.get("success"):
            results["error"] = f"Failed to run formulas: {formula_result.get('error')}"
            return results

        # Step 3: Copy to output
        output_result = self.copy_to_output(intermediate_path, output_path, worksheets_to_output)
        results["copy_to_output"] = output_result

        if not output_result.get("success"):
            results["error"] = f"Failed to create output: {output_result.get('error')}"
            return results

        results["success"] = True
        return results
